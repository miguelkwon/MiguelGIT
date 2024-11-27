import os
import tkinter as tk
import win32gui
import win32com.client
from datetime import datetime, timedelta
import subprocess
import shutil
import psutil
import time
from openpyxl import load_workbook
import pandas as pd
from tkinter import Label
from tkinter import ttk
import calendar
import requests
import json
from tkinter import messagebox
from PIL import Image, ImageTk
import win32con
import threading
import winsound
from tkinter import messagebox, simpledialog
import pygetwindow as gw
import win32api
import openpyxl
import glob
from bs4 import BeautifulSoup
from datetime import datetime, date, timedelta
import pandas as pd
import yfinance as yf
from pystray import Icon, Menu, MenuItem
from tkcalendar import Calendar
from pynput import keyboard
import ctypes



def some_function():
    import time
    time.sleep(1)
    
    
def get_week_info():
    
   # 오늘 날짜
    today = datetime.now()

    # 월과 주 계산
    month_name = today.strftime("%b")  # Nov 등 약어 형태
    week_number = (today.day - 1) // 7 + 1  # 해당 월의 주 계산

    # 날짜 형식: YYYYMMDD
    formatted_date = today.strftime("%Y%m%d")

    # 결과 문자열 생성
    return f"{month_name}. week {week_number}.({formatted_date})"


def get_latest_weekly_report():
    # 기본 폴더 경로
    base_path = r"F:\OneDrive - Radiant Vision Systems\01_RVS\01_Weekly,Activity_Report"
    
    try:
        # 폴더 내 모든 하위 폴더 가져오기
        subfolders = [f for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
        # 최신 폴더 찾기 (YYYYMM 형식)
        latest_folder = max(subfolders, key=lambda x: datetime.strptime(x, "%Y%m"))
        latest_folder_path = os.path.join(base_path, latest_folder)

        # 해당 폴더 내 Weekly report_Korean Team_으로 시작하는 파일 찾기
        files = [f for f in os.listdir(latest_folder_path) 
                 if f.startswith("Weekly report_Korean Team_") and f.endswith(".xlsx")]

        if not files:
            raise FileNotFoundError(f"폴더 '{latest_folder_path}'에 'Weekly report_Korean Team_'으로 시작하는 파일이 없습니다.")
        
        # 가장 최근 파일 찾기 (파일 수정 시간 기준)
        latest_file = max(files, key=lambda x: os.path.getmtime(os.path.join(latest_folder_path, x)))
        latest_file_path = os.path.join(latest_folder_path, latest_file)

        return latest_file_path
    except Exception as e:
        raise FileNotFoundError(f"최신 Weekly report 파일을 찾을 수 없습니다. 오류: {e}")

def open_and_print_latest_report():
    try:
        latest_file_path = get_latest_weekly_report()
        print(f"Opening the latest weekly report: {latest_file_path}")
        
        # 엑셀 파일 읽기
        df = pd.read_excel(latest_file_path)
        print(df)
    except Exception as e:
        print(f"Error opening the latest weekly report: {e}")



def extract_date_from_filename(filename):
    try:
        # 파일 이름 형식: Weekly report_Korean Team_YYYYMMDD_KwangHo Kwon.xlsx
        parts = filename.split("_")  # 파일 이름을 "_"로 분리
        if len(parts) >= 4 and parts[3].isdigit():
            date_part = parts[3]  # 날짜 부분 추출 (예: 20241110)
            return datetime.strptime(date_part, "%Y%m%d")
        else:
            raise ValueError(f"파일 이름이 예상 형식과 다릅니다: {filename}")
    except ValueError as ve:
        raise ValueError(f"파일 이름에서 날짜를 추출할 수 없습니다: {filename}. 오류: {ve}")
    
def get_latest_activity_report():
    # 기본 폴더 경로
    base_path = r"F:\OneDrive - Radiant Vision Systems\01_RVS\01_Weekly,Activity_Report"
    
    try:
        # 폴더 내 모든 하위 폴더 가져오기
        subfolders = [f for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
        # 최신 폴더 찾기 (YYYYMM 형식)
        latest_folder = max(subfolders, key=lambda x: datetime.strptime(x, "%Y%m"))
        latest_folder_path = os.path.join(base_path, latest_folder)

        # 해당 폴더 내 Weekly report_Korean Team_으로 시작하는 파일 찾기
        files = [f for f in os.listdir(latest_folder_path) 
                 if f.startswith("Activity_Report_") and f.endswith(".xlsx")]

        if not files:
            raise FileNotFoundError(f"폴더 '{latest_folder_path}'에 'Activity_Report_'으로 시작하는 파일이 없습니다.")
        
        # 가장 최근 파일 찾기 (파일 수정 시간 기준)
        latest_file = max(files, key=lambda x: os.path.getmtime(os.path.join(latest_folder_path, x)))
        latest_file_path = os.path.join(latest_folder_path, latest_file)

        return latest_file_path
    except Exception as e:
        raise FileNotFoundError(f"최신 Activity_Report 파일을 찾을 수 없습니다. 오류: {e}")

def open_latest_weekly_report():
    try:
        # 최신 파일 경로 가져오기
        latest_file_path = get_latest_weekly_report()
        print(f"최신 파일 경로: {latest_file_path}")

        # 파일 열기
        subprocess.Popen([latest_file_path], shell=True)
        print("파일이 열렸습니다.")
    except FileNotFoundError as e:
        print(e)
    except Exception as e:
        print(f"파일 열기 중 오류 발생: {e}")

def open_latest_activity_report():
    try:
        # 최신 파일 경로 가져오기
        report_path = get_latest_activity_report()
        print(f"최신 파일 경로: {report_path}")
        subprocess.Popen([report_path], shell=True)  # 파일 실행
    except FileNotFoundError as e:
        print(e)

def get_latest_weeklyreport_path():
    import os
    from datetime import datetime

    # 기본 경로
    base_path = r"F:\OneDrive - Radiant Vision Systems\01_RVS\01_Weekly,Activity_Report"

    try:
        # 모든 하위 폴더 가져오기
        subfolders = [f for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
        if not subfolders:
            raise FileNotFoundError("하위 폴더를 찾을 수 없습니다.")

        # 가장 최근 폴더 찾기 (YYYYMM 형식)
        latest_folder = max(subfolders, key=lambda x: datetime.strptime(x, "%Y%m"))
        folder_path = os.path.join(base_path, latest_folder)

        # Weekly report_Korean Team_으로 시작하는 파일 필터링
        files = [f for f in os.listdir(folder_path) if f.startswith("Weekly report_Korean Team_") and f.endswith(".xlsx")]
        if not files:
            raise FileNotFoundError(f"폴더 '{folder_path}'에 'Weekly report_Korean Team_'으로 시작하는 파일이 없습니다.")

        # 가장 최신 파일 찾기 (파일 이름 기준 또는 수정 시간 기준)
        latest_file = max(files, key=lambda f: os.path.getmtime(os.path.join(folder_path, f)))
        file_path = os.path.join(folder_path, latest_file)

        return file_path
    except Exception as e:
        raise FileNotFoundError(f"최신 보고서 파일을 찾는 중 오류 발생: {e}")

def get_latest_activityreport_path():
    import os
    from datetime import datetime

    # 기본 경로
    base_path = r"F:\OneDrive - Radiant Vision Systems\01_RVS\01_Weekly,Activity_Report"

    try:
        # 모든 하위 폴더 가져오기
        subfolders = [f for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
        if not subfolders:
            raise FileNotFoundError("하위 폴더를 찾을 수 없습니다.")

        # 가장 최근 폴더 찾기 (YYYYMM 형식)
        latest_folder = max(subfolders, key=lambda x: datetime.strptime(x, "%Y%m"))
        folder_path = os.path.join(base_path, latest_folder)

        # Weekly report_Korean Team_으로 시작하는 파일 필터링
        files = [f for f in os.listdir(folder_path) if f.startswith("Activity_Report_") and f.endswith(".xlsx")]
        if not files:
            raise FileNotFoundError(f"폴더 '{folder_path}'에 'Activity_Report_'으로 시작하는 파일이 없습니다.")

        # 가장 최신 파일 찾기 (파일 이름 기준 또는 수정 시간 기준)
        latest_file = max(files, key=lambda f: os.path.getmtime(os.path.join(folder_path, f)))
        file_path = os.path.join(folder_path, latest_file)

        return file_path
    except Exception as e:
        raise FileNotFoundError(f"최신 보고서 파일을 찾는 중 오류 발생: {e}")
    
def copy_latest_weekly_report():
    # 기본 경로
    base_path = r"F:\OneDrive - Radiant Vision Systems\01_RVS\01_Weekly,Activity_Report"

    try:
        # 현재 날짜와 관련 정보
        today = datetime.now()
        current_month_folder = today.strftime("%Y%m")  # 예: 202411
        current_date = today.strftime("%Y%m%d")       # 예: 20241103

        # 현재 월의 첫 번째 날
        current_month_start = today.replace(day=1)
        # 이전 월의 마지막 날짜
        previous_month_end = current_month_start - timedelta(days=1)
        previous_month_folder = previous_month_end.strftime("%Y%m")
        previous_date = previous_month_end.strftime("%Y%m%d")  # 예: 20241031

        # 현재 month 폴더 경로 생성
        current_folder_path = os.path.join(base_path, current_month_folder)
        if not os.path.exists(current_folder_path):
            os.makedirs(current_folder_path)
            print(f"새로운 month 폴더 생성: {current_folder_path}")

        # 모든 하위 폴더 가져오기
        subfolders = [f for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
        if not subfolders:
            raise FileNotFoundError("하위 폴더를 찾을 수 없습니다.")

        # 가장 최신 month 폴더 찾기
        latest_month_folder = max(subfolders, key=lambda x: datetime.strptime(x, "%Y%m"))
        latest_folder_path = os.path.join(base_path, latest_month_folder)

        # 최신 month 폴더에서 Weekly report 파일 찾기
        files = [f for f in os.listdir(latest_folder_path) if f.startswith("Weekly report_Korean Team_") and f.endswith(".xlsx")]
        if not files:
            raise FileNotFoundError(f"폴더 '{latest_folder_path}'에 Weekly report 파일이 없습니다.")

        # 최신 파일 찾기
        latest_file = max(files, key=lambda f: os.path.getmtime(os.path.join(latest_folder_path, f)))
        latest_file_path = os.path.join(latest_folder_path, latest_file)

        

        # 두 번째 파일 복사 및 수정 (현재 날짜 기반)
        current_file_name = f"Weekly report_Korean Team_{current_date}_KwangHo Kwon.xlsx"
        current_file_path = os.path.join(current_folder_path, current_file_name)

        if os.path.exists(current_file_path):
            print(f"두 번째 파일이 이미 존재합니다: {current_file_path}")
        else:
            shutil.copy2(latest_file_path, current_file_path)
            print(f"두 번째 파일이 복사되었습니다: {current_file_path}")
            update_monthly_summary(current_file_path)
            print(f"두 번째 파일이 수정되었습니다: {current_file_path}")

        # 두 파일 열기        
        subprocess.Popen([current_file_path], shell=True)
        

    except Exception as e:
        print(f"Weekly report 복사 중 오류 발생: {e}")

def copy_and_update_activity_report():
    # 기본 경로
    base_path = r"F:\OneDrive - Radiant Vision Systems\01_RVS\01_Weekly,Activity_Report"

    try:
        # 현재 날짜와 관련 정보
        today = datetime.now()
        current_month_folder = today.strftime("%Y%m")
        current_date = today.strftime("%Y%m%d")

        # 현재 month 폴더 경로
        current_folder_path = os.path.join(base_path, current_month_folder)
        if not os.path.exists(current_folder_path):
            os.makedirs(current_folder_path)  # 현재 month 폴더가 없으면 생성

        # 모든 하위 폴더 가져오기
        subfolders = [f for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
        if not subfolders:
            raise FileNotFoundError("하위 폴더를 찾을 수 없습니다.")

        # 가장 최근 month 폴더 찾기
        latest_month_folder = max(subfolders, key=lambda x: datetime.strptime(x, "%Y%m"))
        latest_folder_path = os.path.join(base_path, latest_month_folder)

        # 최신 month 폴더에서 Activity_Report 파일 찾기
        files = [f for f in os.listdir(latest_folder_path) if f.startswith("Activity_Report_") and f.endswith(".xlsx")]
        if not files:
            raise FileNotFoundError(f"폴더 '{latest_folder_path}'에 Activity_Report 파일이 없습니다.")

        # 최신 파일 찾기
        latest_file = max(files, key=lambda f: os.path.getmtime(os.path.join(latest_folder_path, f)))
        latest_file_path = os.path.join(latest_folder_path, latest_file)

        # 복사 대상 파일 이름
        new_file_name = f"Activity_Report_{current_date}_KwangHo Kwon.xlsx"
        new_file_path = os.path.join(current_folder_path, new_file_name)

        # 파일이 이미 존재하면 복사를 건너뛰고 파일 열기
        if os.path.exists(new_file_path):
            print(f"파일이 이미 존재합니다. 열기만 수행합니다: {new_file_path}")
        else:
            # 파일 복사
            shutil.copy2(latest_file_path, new_file_path)
            print(f"파일이 복사되었습니다: {new_file_path}")

            # Excel 파일 수정: A7, A14, A21, A28, A35 업데이트
            update_activity_report_dates(new_file_path)
            print(f"Excel 파일이 업데이트되었습니다: {new_file_path}")

        # 파일 열기
        subprocess.Popen([new_file_path], shell=True)
        print(f"파일이 열렸습니다: {new_file_path}")
        
        os.startfile(current_folder_path)

    except Exception as e:
        print(f"Activity_Report 복사 및 수정 중 오류 발생: {e}")

def update_activity_report_dates(file_path):
    # 이번 주의 날짜 계산
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())  # 이번 주 월요일 계산

    # 날짜 리스트 (월~금)
    week_dates = [(monday + timedelta(days=i)).strftime("%m/%d/%Y") for i in range(5)]

    # Excel 파일 열기 및 수정
    workbook = load_workbook(file_path)
    sheet = workbook.active  # 기본 활성화된 시트

    # 셀 업데이트
    cells = ['A7', 'A14', 'A21', 'A28', 'A35']
    for cell, date in zip(cells, week_dates):
        sheet[cell] = date

    # 변경 사항 저장
    workbook.save(file_path)


def update_latest_weekly_report():
    base_path = r"F:\OneDrive - Radiant Vision Systems\01_RVS\01_Weekly,Activity_Report"

    try:
        # 오늘 날짜와 관련 정보
        today = datetime.now()
        current_month_folder = today.strftime("%Y%m")  # 예: 202412
        current_date = today.strftime("%Y%m%d")       # 예: 20241202

        # 오늘 날짜의 월 폴더 경로
        current_folder_path = os.path.join(base_path, current_month_folder)
        if not os.path.exists(current_folder_path):
            os.makedirs(current_folder_path)  # 월 폴더가 없으면 생성
            print(f"새로운 month 폴더 생성: {current_folder_path}")
        else:
            print(f"기존 폴더 존재: {current_folder_path}")

        # 모든 하위 폴더 가져오기
        subfolders = [f for f in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, f))]
        if not subfolders:
            raise FileNotFoundError("하위 폴더를 찾을 수 없습니다.")

        # 가장 최신 month 폴더 찾기
        latest_month_folder = max(subfolders, key=lambda x: datetime.strptime(x, "%Y%m"))
        latest_folder_path = os.path.join(base_path, latest_month_folder)

        # 최신 month 폴더에서 Weekly report 파일 찾기
        files = [f for f in os.listdir(latest_folder_path) if f.startswith("Weekly report_Korean Team_") and f.endswith(".xlsx")]
        if not files:
            raise FileNotFoundError(f"폴더 '{latest_folder_path}'에 Weekly report 파일이 없습니다.")

        # 최신 파일 찾기
        latest_file = max(files, key=lambda f: os.path.getmtime(os.path.join(latest_folder_path, f)))
        latest_file_path = os.path.join(latest_folder_path, latest_file)

        # 저장할 파일 경로
        new_file_name = f"Weekly report_Korean Team_{current_date}_KwangHo Kwon.xlsx"
        new_file_path = os.path.join(current_folder_path, new_file_name)

        # 파일 복사 및 수정
        shutil.copy2(latest_file_path, new_file_path)
        print(f"파일이 복사되었습니다: {new_file_path}")
        update_monthly_summary(new_file_path)
        print(f"파일이 수정되었습니다: {new_file_path}")

        # 파일 열기
        subprocess.Popen([new_file_path], shell=True)
        
        

    except Exception as e:
        print(f"파일 수정 중 오류 발생: {e}")


def update_monthly_summary(file_path):
    try:
        # 이번 달 정보 가져오기
        today = datetime.now()
        month_start = today.replace(day=1)  # 이번 달 1일
        next_month = (month_start + timedelta(days=31)).replace(day=1)
        days_in_month = (next_month - month_start).days  # 이번 달 총 일수

        # 날짜 리스트 생성
        day_list = [(month_start + timedelta(days=i)).strftime("%m/%d/%Y") for i in range(days_in_month)]

        # Excel 파일 열기
        workbook = load_workbook(file_path)

        # Summary 시트 수정
        if "Summary" in workbook.sheetnames:
            summary_sheet = workbook["Summary"]
            for i, date in enumerate(day_list, start=9):
                if i > 38:  # A9부터 A39까지만
                    break
                summary_sheet[f"A{i}"] = date

        # KwangHo Kwon 시트 수정
        if "KwangHo Kwon" in workbook.sheetnames:
            kwangho_sheet = workbook["KwangHo Kwon"]
            for i, date in enumerate(day_list, start=8):
                if i > 37:  # A8부터 A38까지만
                    break
                kwangho_sheet[f"A{i}"] = date

        # 파일 저장
        workbook.save(file_path)
    except Exception as e:
        print(f"Excel 파일 수정 중 오류 발생: {e}")


# Outlook 실행 및 새 메일 쓰기
def is_outlook_running():
    """Outlook 실행 여부 확인"""
    for process in psutil.process_iter(attrs=['name']):
        if process.info['name'] and 'OUTLOOK.EXE' in process.info['name']:
            return True
    return False


def is_onenote_running():
    """Outlook 실행 여부 확인"""
    for process in psutil.process_iter(attrs=['name']):
        if process.info['name'] and 'ONENOTE.EXE' in process.info['name']:
            return True
    return False


def is_truetest_running():
    """TrueTest 실행 여부 확인"""
    for process in psutil.process_iter(attrs=['name']):
        if process.info['name'] and 'TrueTest.EXE' in process.info['name']:
            return True
    return False

def is_winmerge_running():
    """TrueTest 실행 여부 확인"""
    for process in psutil.process_iter(attrs=['name']):
        if process.info['name'] and 'WinMerge.EXE' in process.info['name']:
            return True
    return False

def open_outlook_and_new_mail():
    try:
        # Outlook 실행 여부 확인
        if not is_outlook_running():
            # Outlook 실행
            outlook_path = r"C:\Program Files\Microsoft Office\root\Office16\OUTLOOK.EXE"  # Outlook 경로 확인
            os.startfile(outlook_path)
            print("Outlook 실행 중이 아니어서 실행합니다.")
            time.sleep(5)  # Outlook이 실행될 시간을 줌
        else:
            print("Outlook이 이미 실행 중입니다. 새 메일만 열겠습니다.")

        week_info = get_week_info()
        report_path1 = get_latest_weeklyreport_path()  # 동적으로 파일 경로 가져오기
        report_path2 = get_latest_activityreport_path()  # 동적으로 파일 경로 가져오기

        # 첫 번째 메일 생성 및 전송
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail1 = outlook.CreateItem(0)  # 0은 메일 아이템
        mail1.Subject = f"Weekly report for {week_info}"
        mail1.To = "sun.kim@radiantvs.com"
        mail1.Body = (
            f"Kim Sun Do BJN,\n\nI have attached Weekly report {week_info}.\n"
            f"Please check the attached file.\nThank you.\n\n"
            f"Best regards,\nMiguel\n\n권광호 / 과장\nKwangho, Kwon / Application Engineer\n"
            f"Radiant Vision Systems Korea Limited\n12F, Seokun Tower\n"
            f"178, Pangyoyork-ro, Bundang-gu, Seongnam-si, Gyeonggi-do\n"
            f"O: +82.31.8017.6797 | C: +82.10.9455.5086\nmiguel.kwon@radiantvs.com"
        )
        mail1.Attachments.Add(report_path1)
        mail1.Display()  # 새 메일 창 띄우기

        # 두 번째 메일 생성 및 전송
        mail2 = outlook.CreateItem(0)  # 0은 메일 아이템
        mail2.Subject = f"Activity_Report for {week_info}"
        mail2.To = "jh.song@radiantvs.com"
        mail2.Body = (
            f"Song Jin Hoon JSJN,\n\nI have attached Activity report {week_info}.\n"
            f"Please check the attached file.\nThank you.\n\n"
            f"Best regards,\nMiguel\n\n권광호 / 과장\nKwangho, Kwon / Application Engineer\n"
            f"Radiant Vision Systems Korea Limited\n12F, Seokun Tower\n"
            f"178, Pangyoyork-ro, Bundang-gu, Seongnam-si, Gyeonggi-do\n"
            f"O: +82.31.8017.6797 | C: +82.10.9455.5086\nmiguel.kwon@radiantvs.com"
        )
        mail2.Attachments.Add(report_path2)
        mail2.Display()  # 새 메일 창 띄우기

    except Exception as e:
        print(f"Outlook 실행 중 오류 발생: {e}")

def open_truetest():
    # try:
        
        if not is_truetest_running():
            # Outlook 실행
            truetest_path = r"C:\Program Files\Radiant Vision Systems\TrueTest 1.8\TrueTest.exe"  
            os.startfile(truetest_path)   
            bring_tt_to_front()        
           
        else:
            
            print("Outlook이 이미 실행 중입니다. 새 메일만 열겠습니다.")

        

    # except Exception as e:
    #     print(f"Outlook 실행 중 오류 발생: {e}")
    
def open_imagej():
    
    imagej_path = r"C:\Program Files\ImageJ\ImageJ.exe"  
    os.startfile(imagej_path)

        

    # except Exception as e:
    #     print(f"Outlook 실행 중 오류 발생: {e}")


def open_winmerge():
    try:
        
        if not is_winmerge_running():
            # Outlook 실행
            winmerge_path = r"C:\Program Files\WinMerge\WinMergeU.exe"  
            os.startfile(winmerge_path)           
           
        else:
            print("Outlook이 이미 실행 중입니다. 새 메일만 열겠습니다.")

        

    except Exception as e:
        print(f"Outlook 실행 중 오류 발생: {e}")


def close_truetest():
    try:
        # Windows의 `taskkill` 명령을 사용하여 프로세스 종료
        os.system('taskkill /F /IM TrueTest.exe')
        print("TrueTest.exe가 종료되었습니다.")
    except Exception as e:
        print(f"TrueTest 종료 중 오류 발생: {e}")






# 실행



def execute_both_reports():     
    open_latest_weekly_report()
    open_latest_activity_report()

def copy_both_reports():     
    copy_and_update_activity_report()
    copy_latest_weekly_report()




# def create_gui():
#     # Tkinter 윈도우 생성
#     root = tk.Tk()
#     root.title("업무 효율 프로그램")
#     root.geometry("400x400")  # 창 크기 설정

    

#     # 버튼 프레임 생성
#     button_frame = tk.Frame(root)
#     button_frame.pack(pady=10)

#     # 1번 버튼: Weekly, Activity 수정
#     btn_1 = tk.Button(
#         button_frame, 
#         text="Weekly, Activity 수정", 
#         width=20, 
#         height=2, 
#         bg="lightgreen",
#         command=copy_both_reports  # 최신 파일 실행 함수 연결
#     )
#     btn_1.grid(row=0, column=0, padx=10, pady=10)  # 1번 버튼 배치

#     # 2번 버튼: TrueTest 종료
#     btn_2 = tk.Button(
#         button_frame, 
#         text="TrueTest 종료", 
#         width=20, 
#         height=2, 
#         bg="lightblue",
#         command=close_truetest  # 최신 파일 실행 함수 연결
#     )
#     btn_2.grid(row=0, column=1, padx=10, pady=10)  # 2번 버튼 배치

#     # 3번 버튼: TrueTest
#     btn_3 = tk.Button(
#         button_frame, 
#         text="TrueTest", 
#         width=20, 
#         height=2, 
#         bg="lightblue",
#         command=open_truetest  # 최신 파일 실행 함수 연결
#     )
#     btn_3.grid(row=1, column=0, padx=10, pady=10)  # 3번 버튼 배치

#     # 4번 버튼: TrueTest 종료
#     btn_4 = tk.Button(
#         button_frame, 
#         text="TrueTest 종료", 
#         width=20, 
#         height=2, 
#         bg="lightblue",
#         command=close_truetest  # 최신 파일 실행 함수 연결
#     )
#     btn_4.grid(row=1, column=1, padx=10, pady=10)  # 4번 버튼 배치

#     # 5번 버튼: WinMerge
#     btn_5 = tk.Button(
#         button_frame, 
#         text="WinMerge", 
#         width=20, 
#         height=2, 
#         bg="lightblue",
#         command=open_winmerge  # 최신 파일 실행 함수 연결
#     )
#     btn_5.grid(row=2, column=0, padx=10, pady=10)  # 5번 버튼 배치

#     # 6번 버튼: WinMerge
#     btn_6 = tk.Button(
#         button_frame, 
#         text="WinMerge", 
#         width=20, 
#         height=2, 
#         bg="lightblue",
#         command=open_winmerge  # 최신 파일 실행 함수 연결
#     )
#     btn_6.grid(row=2, column=1, padx=10, pady=10)  # 6번 버튼 배치

#     # 9번 버튼: Weekly, Activity 수정
#     btn_9 = tk.Button(
#         button_frame, 
#         text="Weekly, Activity 수정", 
#         width=20, 
#         height=2, 
#         bg="orange",
#         command=copy_both_reports  # 최신 파일 실행 함수 연결
#     )
#     btn_9.grid(row=3, column=0, padx=10, pady=10)  # 9번 버튼 배치

#     # 10번 버튼: Weekly, Activity 보내기
#     btn_10 = tk.Button(
#         button_frame, 
#         text="Weekly, Activity 보내기", 
#         width=20, 
#         height=2, 
#         bg="orange",
#         command=open_outlook_and_new_mail  # 최신 파일 실행 함수 연결
#     )
#     btn_10.grid(row=3, column=1, padx=10, pady=10)  # 10번 버튼 배치

#     # 프로그램 종료 버튼
#     exit_button = tk.Button(root, text="종료", command=root.quit, bg="red", fg="white")
#     exit_button.pack(pady=20)

#     # Tkinter 메인 루프 실행
#     root.mainloop()

# # 여기서 copy_both_reports, close_truetest, open_truetest, open_winmerge, open_outlook_and_new_mail 함수를 정의해야 합니다.



# import tkinter as tk
# 

# def copy_both_reports():
#     print("Weekly, Activity 수정 실행")

# def close_truetest():
#     print("TrueTest 종료 실행")

# def open_truetest():
#     print("TrueTest 실행")

# def open_winmerge():
#     print("WinMerge 실행")

# def open_outlook_and_new_mail():
#     print("Weekly, Activity 보내기 실행")

# import tkinter as tk
# from tkinter import ttk



# def copy_both_reports():
#     print("Weekly, Activity 수정 실행")

# def close_truetest():
#     print("TrueTest 종료 실행")

# def open_truetest():
#     print("TrueTest 실행")

# def open_winmerge():
#     print("WinMerge 실행")

# def open_outlook_and_new_mail():
#     print("Weekly, Activity 보내기 실행")

def update_time(label, root):
    """현재 시간과 요일을 업데이트 (요일 약어)"""
    current_time = datetime.now()
    # 요일을 약어로 가져오기 (Mon, Tue, Wed, ...)
    weekday = current_time.strftime("%a")
    # 시간 문자열에 요일 약어 추가
    time_with_weekday = current_time.strftime(f"%Y-%m-%d-{weekday}-%H:%M:%S ")
    label.config(text=time_with_weekday)
    label.after(1000, update_time, label, root)
    
    label.bind("<Button-1>", lambda event: open_alarm_popup(root))
    label.bind("<Button-3>", lambda event: copy_time_to_clipboard(time_with_weekday))


def copy_time_to_clipboard(text):
    """현재 시간을 클립보드에 복사"""
    root = tk.Tk()
    root.withdraw()  # GUI 숨기기
    root.clipboard_clear()  # 기존 클립보드 내용 삭제
    root.clipboard_append(text)  # 텍스트 복사
    root.update()  # 클립보드 업데이트
    root.destroy()  # Tkinter 윈도우 닫기
    # messagebox.showinfo("Copied", "Current time has been copied to clipboard!")


# 파일 경로
D_DAY_FILE = "d_day.json"
    
# D-Day 저장 함수
def save_d_day(selected_date):
        with open(D_DAY_FILE, "w") as file:
            json.dump({"d_day": selected_date}, file)

# D-Day 읽기 함수
def load_d_day():
    if os.path.exists(D_DAY_FILE):
        with open(D_DAY_FILE, "r") as file:
            data = json.load(file)
            return data.get("d_day")
    return None



    

 # D-Day 계산 함수
def calculate_d_day(selected_date, label, update_callback):
    # 기준 날짜를 2024년 11월 22일로 설정
    today = datetime(2024, 11, 22)
    selected_date_obj = datetime.strptime(selected_date, "%Y-%m-%d")
    delta = (selected_date_obj - today).days

    if delta == 0:
        d_day_text = "D-Day: Today!"
    elif delta == 1:  # 내일이 D-Day인 경우
        d_day_text = "D-Day: 1   "
    elif delta > 0:
        d_day_text = f"D-Day: {delta}   "
    else:
        d_day_text = f"D-Day: {-delta}   "

    # D-Day 텍스트를 update_callback을 통해 업데이트
    update_callback(d_day_text)
    save_d_day(selected_date)



    # D-Day 취소 함수
def cancel_d_day(update_callback):
    """D-Day 취소 시 D-Day 텍스트를 'Not Set'으로 설정"""
    update_callback("D-Day: Not Set")  # 텍스트 업데이트
    if os.path.exists(D_DAY_FILE):
        os.remove(D_DAY_FILE)  # D-Day 데이터 파일 삭제


# 달력 선택 팝업 함수
def open_calendar(label, update_callback):
    def set_date():
        selected_date = calendar.get_date()
        top.destroy()
        calculate_d_day(selected_date, label, update_callback)

    top = tk.Toplevel(root)
    top.title("Select a Date")

    calendar = Calendar(top, date_pattern="y-mm-dd")
    calendar.pack(padx=10, pady=10)

    select_btn = ttk.Button(top, text="Select", command=set_date)
    select_btn.pack(pady=5)

class CalendarApp:
    def __init__(self, parent):
        self.parent = parent
        self.current_year = datetime.now().year
        self.current_month = datetime.now().month

        # 초기 UI 생성
        self.create_ui()

    def create_ui(self):
        """달력 UI 생성"""
        self.cal_frame = ttk.Frame(self.parent)
        self.cal_frame.pack(pady=0)

        # 달력 표시
        self.display_calendar()
      
      
    
            
        
    def display_calendar(self):
        """현재 달력을 고정된 프레임 내부에 표시"""
        # 기존 위젯 삭제
        for widget in self.cal_frame.winfo_children():
            widget.destroy()

        # 다크 모드 스타일 설정
        style = ttk.Style()
        style.configure("Dark.TFrame", background="#121212")  # 프레임 배경 검은색
        style.configure("Dark.TLabel", background="#121212", foreground="#FFFFFF")  # 라벨 배경 검은색, 텍스트 흰색

        # 스크롤 가능한 Canvas 생성
        canvas = tk.Canvas(self.cal_frame, width=400, height=420, bg="#121212", highlightthickness=0)  # 검은 배경
        canvas.grid(row=0, column=0, sticky="nsew")

        # 스크롤바 추가
        scrollbar = ttk.Scrollbar(self.cal_frame, orient="vertical", command=canvas.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Canvas와 Scrollbar 연결
        canvas.configure(yscrollcommand=scrollbar.set)

        # Canvas 내부에 Frame 생성
        scrollable_frame = ttk.Frame(canvas, style="Dark.TFrame")
        scrollable_frame_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        # Canvas 크기 변경 시 내부 프레임 크기 갱신
        def update_scrollregion(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        scrollable_frame.bind("<Configure>", update_scrollregion)

        now = datetime(self.current_year, self.current_month, 1)
        today = datetime.now()

        # 현재 연도의 몇 주차인지 계산
        current_week_number = today.isocalendar()[1]

        # 영어로 월 이름 설정
        month_name = now.strftime("%B")  # 예: "January"
        week_day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

        # 헤더 프레임 (년/월과 화살표 버튼)
        header_frame = ttk.Frame(scrollable_frame, style="Dark.TFrame")
        header_frame.grid(row=0, column=0, columnspan=7, pady=10)

        # 이전 달 버튼
        prev_button = ttk.Button(
            header_frame,
            text="←",
            command=self.show_prev_month,
            width=3,
        )
        prev_button.grid(row=0, column=0, padx=5)

        # 현재 연도와 월 텍스트
        header_label = ttk.Label(
            header_frame,
            text=f"{month_name} {self.current_year}",  # 예: "January 2024"
            font=("Arial", 14, "bold"),
            style="Dark.TLabel",  # 다크 모드 스타일 사용
        )
        header_label.grid(row=0, column=1, padx=10)

        # 다음 달 버튼
        next_button = ttk.Button(
            header_frame,
            text="→",
            command=self.show_next_month,
            width=3,
        )
        next_button.grid(row=0, column=2, padx=5)
        
        
        # D-Day와 주차를 함께 표시할 Label 생성
        combined_label = ttk.Label(
            scrollable_frame,
            text="D-Day: Not Set  | Current Week: Calculating...",
            font=("Arial", 12, "bold"),
            style="Dark.TLabel"
        )
        combined_label.grid(row=1, column=0, columnspan=7, pady=0)

        # D-Day 설정 함수와 이벤트
        def update_combined_label(d_day_text=None):
            """D-Day 텍스트와 주차를 결합하여 라벨 업데이트"""
            d_day_text = d_day_text if d_day_text else "D-Day: Not Set"
            current_week = datetime.now().isocalendar()[1]
            combined_label.config(text=f"{d_day_text} |    Week: {current_week}")

        # 좌클릭으로 D-Day 설정
        combined_label.bind("<Button-1>", lambda event: open_calendar(combined_label, update_combined_label))
        # 우클릭으로 D-Day 취소
        combined_label.bind("<Button-3>", lambda event: cancel_d_day(update_combined_label))

        # 저장된 D-Day 로드
        saved_date = load_d_day()
        if saved_date:
            calculate_d_day(saved_date, combined_label, update_combined_label)
        else:
            update_combined_label()

        for i, day in enumerate(week_day_names):
            color = "#DDDDDD" if i < 5 else "#FF6666"  # 평일: 밝은 회색, 주말: 빨간색
            lbl = tk.Label(
                scrollable_frame,
                text=day,
                font=("Arial", 12, "bold"),
                anchor="center",
                width=4,
                height=2,
                relief="flat",  # 테두리 제거
                bd=0,  # 테두리 두께 제거
                background="#121212",  # 어두운 배경색
                foreground=color,  # 텍스트 색상
            )
            lbl.grid(row=2, column=i, padx=5, pady=5)

        # 날짜 출력
        cal = calendar.Calendar()
        weeks = cal.monthdayscalendar(self.current_year, self.current_month)
        row = 3

        for week_index, week in enumerate(weeks):
            for col, day in enumerate(week):
                if day == 0:  # 빈 날짜는 표시하지 않음
                    lbl = tk.Label(
                        scrollable_frame,
                        text="",
                        font=("Arial", 10),
                        width=4,
                        height=2,
                        relief="flat",  # 테두리 제거
                        bd=0,  # 테두리 두께 제거
                        background="#121212",  # 어두운 배경
                    )
                else:
                    # 오늘 날짜 강조
                    if today.year == self.current_year and today.month == self.current_month and day == today.day:
                        lbl = tk.Label(
                            scrollable_frame,
                            text=day,
                            font=("Arial", 10, "bold"),
                            bg="#FF6666",  # 강조 색상
                            relief="flat",  # 테두리 제거
                            bd=0,  # 테두리 두께 제거
                            width=4,
                            height=2,
                        )
                    else:
                        lbl = tk.Label(
                            scrollable_frame,
                            text=day,
                            font=("Arial", 10),
                            relief="flat",  # 테두리 제거
                            bd=0,  # 테두리 두께 제거
                            background="#121212",  # 어두운 배경
                            foreground="#FFFFFF",  # 텍스트 색상
                            width=4,
                            height=2,
                        )

                    # 날짜 더블 클릭 이벤트 바인딩
                    lbl.bind(
                        "<Double-1>",
                        lambda event, selected_day=day: self.open_outlook_calendar(
                            datetime(self.current_year, self.current_month, selected_day)
                        ),
                    )
                lbl.grid(row=row, column=col, padx=5, pady=5)
            row += 1








    def show_prev_month(self):
        """이전 달로 이동"""
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self.display_calendar()

    def show_next_month(self):
        """다음 달로 이동"""
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self.display_calendar()

    def open_outlook_calendar(self, date):
        try:
            # Outlook 프로세스 강제 초기화
            # os.system("taskkill /F /IM outlook.exe")  # 실행 중인 Outlook 프로세스 종료
            outlook = win32com.client.Dispatch("Outlook.Application")  # Outlook 실행
            namespace = outlook.GetNamespace("MAPI")

            # 캘린더 이동
            calendar_folder = namespace.GetDefaultFolder(9)  # 9은 캘린더 폴더
            explorer = calendar_folder.GetExplorer()
            explorer.Display()

            # 창을 전체 크기로 설정
            explorer.WindowState = 1  # 1: 최대화 (Outlook 창 설정)

            # 날짜 설정 (날짜와 시간을 UTC로 정확히 설정)
            explorer.CurrentView.GoToDate(date)

            # 새 약속 작성
            appointment = outlook.CreateItem(1)  # 1은 AppointmentItem
            appointment.Start = date.replace(hour=9, minute=0, second=0)  # 오전 9시로 설정
            appointment.Subject = ""  # 기본 제목 설정
            appointment.Body = "Details about this appointment..."  # 기본 내용 설정
            appointment.Display()  # 약속 창 열기

            # 약간의 대기 시간 (Outlook 창 로드 완료)
            time.sleep(1)

            # Outlook 창을 맨 앞에 표시
            def enum_window_callback(hwnd, wildcard):
                if win32gui.IsWindowVisible(hwnd) and wildcard in win32gui.GetWindowText(hwnd):
                    win32gui.SetForegroundWindow(hwnd)

            # "Outlook"이 포함된 창을 찾아서 활성화
            win32gui.EnumWindows(enum_window_callback, "Outlook")

        except Exception as e:
            print(f"Outlook 열기에 실패했습니다: {e}")






# def create_gui():
#     # Tkinter 윈도우 생성
#     root = tk.Tk()
#     root.title("업무 효율 프로그램")

#     # 화면 크기와 위치 계산
#     window_width = 400
#     window_height = 1000

#     # 화면의 전체 크기를 가져오기
#     screen_width = root.winfo_screenwidth()
#     screen_height = root.winfo_screenheight()

#     # 창 위치 계산 (오른쪽 끝 상단)
#     x_pos = screen_width - window_width  # 오른쪽 끝
#     y_pos = 0  # 상단

#     # 창 크기 및 위치 설정
#     root.geometry(f"{window_width}x{window_height}+{x_pos}+{y_pos}")
#     root.configure(bg="#f8f8f8")  # 배경색 설정

#     # 현재 시간 표시 라벨
#     title_label = ttk.Label(root, font=("Arial", 20, "bold"))
#     title_label.pack(pady=20)
#     update_time(title_label)  # 시간 업데이트 시작

#     # 현재 달력 표시
#     CalendarApp(root)

#     # 버튼 프레임 생성
#     button_frame = ttk.Frame(root, padding="0")
#     button_frame.pack(pady=20)

#     # 버튼 스타일 생성
#     style = ttk.Style()
#     style.configure("TButton", font=("Arial", 12), padding=10)
#     style.map("TButton", background=[("active", "#d7f3e3")])

#     # 버튼 목록 정의
#     buttons = [
#         ("TrueTest 실행", open_truetest),
#         ("TrueTest 종료", close_truetest),
#         ("WinMerge 실행", open_winmerge),
#         ("Weekly, Activity 수정", copy_both_reports),
#         ("Weekly, Activity 보내기", open_outlook_and_new_mail),
#     ]

#     # 버튼 생성 및 배치
#     for i, (text, command) in enumerate(buttons):
#         button = ttk.Button(button_frame, text=text, command=command)
#         button.grid(row=i // 2, column=i % 2, padx=10, pady=10)
    

    


#     # Tkinter 메인 루프 실행
#     root.mainloop()

def open_outlook():
    """Outlook 실행"""
    try:
        
        if not is_outlook_running():
            # Outlook 실행
            outlook_path = r"C:/Program Files/Microsoft Office/root/Office16/OUTLOOK.EXE"  
            os.startfile(outlook_path)           
           
        else:
            # messagebox.showinfo("Information", "Outlook이 이미 실행 중입니다.")
            bring_outlook_to_front()
    except Exception as e:
            print(f"Outlook 열기에 실패했습니다: {e}")

def open_onenote():
    """OneNote 실행"""
    try:
        if not is_onenote_running():
            onenote_path = r"C:/Program Files/Microsoft Office/root/Office16/ONENOTE.EXE"  # OneNote 실행 파일 경로
            os.startfile(onenote_path)
        else:
            # messagebox.showinfo("Information", "Outlook이 이미 실행 중입니다.")
            bring_onenote_to_front()
    except Exception as e:
        messagebox.showerror("Error", f"OneNote 실행 실패: {e}")
        

def open_teams():
    """OneNote 실행"""
    
    teams_path = r"C:\Users\ssa2p\AppData\Local\Microsoft\WindowsApps\ms-teams.exe"  # OneNote 실행 파일 경로
    os.startfile(teams_path)
    

def open_kt():
        
    kt_path = r"C:\Program Files (x86)\Kakao\KakaoTalk\KakaoTalk.exe"  # OneNote 실행 파일 경로
    os.startfile(kt_path)



def bring_outlook_to_front():
    """이미 실행 중인 Outlook 창을 최상위로 가져오기"""
    try:
        # Outlook 창 핸들 가져오기
        hwnd = win32gui.FindWindow(None, "Inbox - miguel.kwon@radiantvs.com - Outlook")  # Outlook 창 제목의 일부를 입력
        if hwnd:
            # 창을 최상위로 가져오고 활성화
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            win32gui.SetForegroundWindow(hwnd)
        else:
            messagebox.showinfo("Information", "Outlook 창을 찾을 수 없습니다.")
    except Exception as e:
        print(f"Outlook 창 활성화 실패: {e}")
        
def bring_tt_to_front():
    """TrueTest 메인 창을 앞으로 가져오기"""
    try:
        # 'TrueTest v1.8' 패턴이 포함된 창 검색
        hwnd = find_window_with_partial_title("TrueTest v1.8")
        if hwnd and win32gui.IsWindow(hwnd):  # 창 핸들 유효성 확인
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)  # 최소화된 창 복원
            win32gui.SetForegroundWindow(hwnd)  # 창을 활성화
            print("TrueTest 메인 창을 앞으로 가져왔습니다.")
        else:
            print("TrueTest 메인 창을 찾을 수 없습니다.")  # 메시지 박스 대신 로그 출력
    except Exception as e:
        print(f"TrueTest 메인 창 활성화 실패: {e}")







def find_window_with_partial_title(partial_title):
    """특정 제목 패턴을 포함한 창 검색"""
    def callback(hwnd, results):
        title = win32gui.GetWindowText(hwnd)
        if partial_title in title:
            results.append((hwnd, title))  # 창 핸들과 제목 추가

    windows = []
    win32gui.EnumWindows(callback, windows)

    # 검색된 창 목록 출력 (디버깅 용도)
    for hwnd, title in windows:
        print(f"Found window: Handle={hwnd}, Title='{title}'")

    # 'TrueTest v1.8'으로 시작하는 창 반환
    for hwnd, title in windows:
        if title.startswith("TrueTest v1.8"):
            return hwnd

    # 첫 번째 검색된 창 반환 (fallback)
    return windows[0][0] if windows else None



def bring_onenote_to_front():
    """이미 실행 중인 Outlook 창을 최상위로 가져오기"""
    try:
        # Outlook 창 핸들 가져오기
        hwnd = find_window_with_partial_title("- OneNote")  # Outlook 창 제목의 일부를 입력
        if hwnd:
            # 창을 최상위로 가져오고 활성화
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
            win32gui.SetForegroundWindow(hwnd)
        else:
            messagebox.showinfo("Information", "Outlook 창을 찾을 수 없습니다.")
    except Exception as e:
        print(f"Outlook 창 활성화 실패: {e}")

# def update_time(label):
#     """현재 시간을 업데이트"""
#     current_time = datetime.now().strftime("%H:%M:%S")
#     label.config(text=current_time)
#     label.after(1000, update_time, label)

def open_alarm_popup(root):
    """알람 설정 창 열기"""
    popup = tk.Toplevel(root)
    popup.title("Set Alarm")
    popup.geometry("300x250")
    popup.configure(bg="#F0F0F0")

    # 현재 시간 표시 및 조정 UI
    current_time = datetime.now()

    time_frame = ttk.Frame(popup)
    time_frame.pack(pady=10)

    # 시간과 분을 표시하는 변수
    hour_var = tk.IntVar(value=current_time.hour)
    minute_var = tk.IntVar(value=current_time.minute)

    # 시간 레이블
    ttk.Label(time_frame, text="Set Alarm Time:", font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=3, pady=5)

    # 시간 조정 버튼과 레이블
    ttk.Button(time_frame, text="▲", command=lambda: hour_var.set((hour_var.get() + 1) % 24)).grid(row=1, column=0)
    ttk.Label(time_frame, textvariable=hour_var, font=("Arial", 16)).grid(row=2, column=0)
    ttk.Button(time_frame, text="▼", command=lambda: hour_var.set((hour_var.get() - 1) % 24)).grid(row=3, column=0)

    ttk.Label(time_frame, text=":", font=("Arial", 16, "bold")).grid(row=2, column=1)

    ttk.Button(time_frame, text="▲", command=lambda: minute_var.set((minute_var.get() + 1) % 60)).grid(row=1, column=2)
    ttk.Label(time_frame, textvariable=minute_var, font=("Arial", 16)).grid(row=2, column=2)
    ttk.Button(time_frame, text="▼", command=lambda: minute_var.set((minute_var.get() - 1) % 60)).grid(row=3, column=2)

    # 알람 설정 버튼
    def set_alarm():
        alarm_time = current_time.replace(hour=hour_var.get(), minute=minute_var.get(), second=0)
        messagebox.showinfo("Alarm Set", f"Alarm set for {alarm_time.strftime('%H:%M')}")
        popup.destroy()
        start_alarm_thread(alarm_time)

    ttk.Button(popup, text="Set Alarm", command=set_alarm).pack(pady=20)

    # 추가 알람 옵션 (빠른 설정)
    ttk.Label(popup, text="Quick Select:", font=("Arial", 12)).pack(pady=5)

    options_frame = ttk.Frame(popup)
    options_frame.pack()

    def quick_set(minutes):
        """빠른 알람 설정"""
        alarm_time = current_time + timedelta(minutes=minutes)
        messagebox.showinfo("Alarm Set", f"Alarm set for {alarm_time.strftime('%H:%M')}")
        popup.destroy()
        start_alarm_thread(alarm_time)

    options = [("30 Min", 30), ("1 Hour", 60), ("2 Hours", 120), ("3 Hours", 180), ("4 Hours", 240)]
    for text, minutes in options:
        ttk.Button(options_frame, text=text, command=lambda m=minutes: quick_set(m)).pack(side="left", padx=5)


def start_alarm_thread(alarm_time):
    """알람 체크를 별도 스레드에서 실행"""
    def check_alarm():
        while True:
            if datetime.now() >= alarm_time:
                show_alarm()
                break

    alarm_thread = threading.Thread(target=check_alarm)
    alarm_thread.daemon = True  # 메인 프로그램 종료 시 스레드도 종료
    alarm_thread.start()


# 소리 중단 플래그
alarm_stop_flag = threading.Event()

def show_alarm():
    """알람 알림 창"""
    # 경고음을 5초 동안 재생
    def play_alarm_sound():
        while not alarm_stop_flag.is_set():  # 플래그가 설정되지 않은 동안 재생
            winsound.Beep(1000, 1000)  # 1000Hz 주파수, 1초 지속
            if alarm_stop_flag.is_set():
                break

    threading.Thread(target=play_alarm_sound, daemon=True).start()  # 비동기로 실행

    # 팝업 창 생성
    alarm_popup = tk.Toplevel(root)
    alarm_popup.title("Popup")

    # 창 크기
    window_width = 400
    window_height = 200

    # 화면 크기
    screen_width = alarm_popup.winfo_screenwidth()
    screen_height = alarm_popup.winfo_screenheight()

    # 화면 가운데 위치 계산
    x_pos = (screen_width // 2) - (window_width // 2)
    y_pos = (screen_height // 2) - (window_height // 2)

    # 창 크기 및 위치 설정
    alarm_popup.geometry(f"{window_width}x{window_height}+{x_pos}+{y_pos}")

    # 팝업 내용
    label = tk.Label(alarm_popup, text="This is a popup!", font=("Arial", 14))
    label.pack(pady=20)

    close_button = tk.Button(alarm_popup, text="Close", command=alarm_popup.destroy)
    close_button.pack(pady=10)
    
    
    # 알람 메시지
    ttk.Label(
        alarm_popup,
        text="Time's up! Your alarm is ringing!",
        font=("Arial", 16, "bold"),
        background="#FFDDDD",
        foreground="#000000",
    ).pack(expand=True, pady=20)

    # 알람 종료 버튼
    def stop_alarm():
        alarm_stop_flag.set()  # 소리 중단 플래그 설정
        alarm_popup.destroy()  # 알람 창 닫기

    ttk.Button(
        alarm_popup,
        text="OK",
        command=stop_alarm,
        style="Alarm.TButton"
    ).pack(pady=10)

    # 스타일 설정 (버튼)
    style = ttk.Style()
    style.configure(
        "Alarm.TButton",
        font=("Arial", 12, "bold"),
        padding=10,
        background="#FF5555",
        foreground="#FFFFFF",
        relief="raised"
    )
def make_folder():
    """오늘 날짜 폴더와 텍스트 파일 생성 및 폴더 열기"""
    base_path = r"F:\01_RVS\05_Miguel"
    today_date = datetime.now().strftime("%Y%m%d")  # 오늘 날짜 (YYYYMMDD 형식)
    folder_path = os.path.join(base_path, today_date)
    file_path = os.path.join(folder_path, f"{today_date}.txt")
    
    # 폴더 생성
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)  # 폴더 생성
        print(f"폴더 생성 완료: {folder_path}")
    else:
        print(f"폴더가 이미 존재합니다: {folder_path}")
    
    # 파일 생성
    if not os.path.exists(file_path):
        with open(file_path, "w", encoding="utf-8") as file:
            file.write(f"파일 생성 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        print(f"파일 생성 완료: {file_path}")
    else:
        print(f"파일이 이미 존재합니다: {file_path}")
    
    # 폴더 열기
    subprocess.Popen(f'explorer "{folder_path}"')  # Windows 탐색기로 폴더 열기
    os.startfile(file_path)

def make_python():
    """example.py 파일 생성 및 기본 프로그램으로 열기, 폴더 열기"""
    base_path = r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python"
    base_filename = "example"
    extension = ".py"
    
    # 기본 파일 이름 생성
    filename = f"{base_filename}{extension}"
    file_path = os.path.join(base_path, filename)
    
    # 숫자를 붙여서 고유한 파일 이름 생성
    counter = 1
    while os.path.exists(file_path):  # 파일이 존재하면 숫자를 증가시킴
        filename = f"{base_filename}{counter}{extension}"
        file_path = os.path.join(base_path, filename)
        counter += 1
    
    # 파일 생성
    with open(file_path, "w", encoding="utf-8") as file:
        file.write("# Example Python file\n")
        print(f"파일 생성 완료: {file_path}")
    
    # 파일을 Windows 기본 설정 프로그램으로 열기
    try:
        os.startfile(file_path)  # 기본 프로그램으로 실행
        print(f"Windows 기본 프로그램으로 파일 열기: {file_path}")
    except OSError as e:
        print(f"파일을 열 수 없습니다: {e}")
    
    # 폴더를 Windows 탐색기로 열기
    try:
        os.startfile(base_path)  # 폴더 열기
        print(f"Windows 탐색기로 폴더 열기: {base_path}")
    except OSError as e:
        print(f"폴더를 열 수 없습니다: {e}")



def load_icon_with_black_bg(image_path, size):
    """이미지 로드 및 배경 검은색 설정"""
    image = Image.open(image_path).convert("RGBA")  # RGBA 모드로 변환
    black_bg = Image.new("RGBA", image.size, (0, 0, 0, 255))  # 검은색 배경 생성
    black_bg.paste(image, (0, 0), image)  # 이미지의 투명 부분에 검은색 배경 추가
    resized_image = black_bg.resize(size, Image.LANCZOS)  # 크기 조정
    return ImageTk.PhotoImage(resized_image)

def center_window(window, width, height):
    """
    주어진 창을 화면 중앙에 배치합니다.
    :param window: Tkinter Toplevel 또는 Tk 객체
    :param width: 창의 너비
    :param height: 창의 높이
    """
    # 화면 너비와 높이 가져오기
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    # 중앙 위치 계산
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2

    # 창 위치 및 크기 설정
    window.geometry(f"{width}x{height}+{x}+{y}")

def create_folders(root):
    # 기본 저장 경로
    default_path = r"F:\01_RVS\05_Miguel"  # 기본 폴더 경로 수정 가능

    # 팝업창 - 기본 경로 확인
    selected_path = simpledialog.askstring(
        "폴더 저장 경로",
        f"폴더를 저장할 경로를 입력하세요 (기본 경로: {default_path}):",
        initialvalue=default_path,
    )

    # 사용자가 취소를 누르면 종료
    if not selected_path:
        messagebox.showinfo("취소", "폴더 생성이 취소되었습니다.")
        return

    # 경로가 없으면 생성
    if not os.path.exists(selected_path):
        os.makedirs(selected_path)
        messagebox.showinfo("폴더 생성", f"새로운 경로가 생성되었습니다: {selected_path}")

    # 팝업창 - 몇 개의 폴더를 만들지 입력
    try:
        folder_count = simpledialog.askinteger("폴더 개수", "만들 폴더 개수를 입력하세요:", minvalue=1)
        if folder_count is None:
            messagebox.showinfo("취소", "폴더 생성이 취소되었습니다.")
            return
    except ValueError:
        messagebox.showerror("입력 오류", "숫자를 입력해야 합니다.")
        return
    
    current_date = datetime.now().strftime("%Y%m%d")
    
    # 폴더 이름 입력 UI 생성
    def collect_folder_names():
        folder_names = [entry.get() for entry in entries]
        if all(folder_names):
            # 폴더 생성
            created_folders = []
            for folder_name in folder_names:
                folder_path = os.path.join(selected_path, folder_name)
                if not os.path.exists(folder_path):
                    os.makedirs(folder_path)
                    created_folders.append(folder_path)

            # 완료 메시지
            if created_folders:
                messagebox.showinfo(
                    "완료",
                    f"다음 폴더가 생성되었습니다:\n{', '.join(created_folders)}",
                    parent=folder_name_window
                )
            else:
                messagebox.showwarning(
                    "주의",
                    "폴더가 이미 존재합니다. 새로운 폴더를 생성하지 않았습니다.",
                    parent=folder_name_window
                )
            os.startfile(selected_path)
            folder_name_window.destroy()
        else:
            messagebox.showerror("입력 오류", "모든 폴더 이름을 입력해야 합니다.", parent=folder_name_window)

    # 새 창 생성
    folder_name_window = tk.Toplevel(root)
    folder_name_window.title("폴더 이름 입력")

    # 창 크기와 위치 설정 (화면 중앙)
    center_window(folder_name_window, 400, 300)

    ttk.Label(folder_name_window, text="폴더 이름을 입력하세요:", font=("Arial", 12)).pack(pady=10)

    # 체크박스 - 날짜 포함 여부
    include_date_var = tk.BooleanVar(value=True)
    ttk.Checkbutton(folder_name_window, text="현재 날짜 포함", variable=include_date_var).pack(pady=5)

    entries = []
    for i in range(folder_count):
        frame = ttk.Frame(folder_name_window)
        frame.pack(fill=tk.X, pady=5)

        # 기본 이름에 날짜 포함 여부 반영
        default_name = f"{current_date}_" if include_date_var.get() else ""

        ttk.Label(frame, text=f"폴더 {i + 1}:", width=10).pack(side=tk.LEFT, padx=5)
        entry = ttk.Entry(frame)
        entry.insert(0, default_name)  # 기본 폴더 이름 설정
        entry.pack(fill=tk.X, expand=True, padx=5)
        entries.append(entry)

    def update_entries():
        for entry in entries:
            if include_date_var.get():
                if not entry.get().startswith(current_date):
                    entry.insert(0, f"{current_date}_")
            else:
                if entry.get().startswith(current_date):
                    entry.delete(0, len(current_date) + 1)

    include_date_var.trace("w", lambda *args: update_entries())

    ttk.Button(folder_name_window, text="확인", command=collect_folder_names).pack(pady=10)

 
def arrange_folders():
    # 현재 열려 있는 모든 창 가져오기
    open_windows = [win for win in gw.getWindowsWithTitle("") if "File Explorer" in win.title]
    
    folder_count = len(open_windows)
    
    time.sleep(1)

    if folder_count == 0:
        messagebox.showinfo("알림", "현재 열려 있는 파일 탐색기가 없습니다.")
        return

    screen_width = win32api.GetSystemMetrics(0)
    screen_height = win32api.GetSystemMetrics(1)

    # 창 위치 설정 함수
    def set_window_position(hwnd, x, y, width, height):
        win32gui.SetWindowPos(hwnd, win32con.HWND_TOP, x, y, width, height, 0)

    # 열려 있는 폴더 수에 따라 창 배치
    if folder_count == 1:
        win = open_windows[0]
        hwnd = win._hWnd
        set_window_position(hwnd, 0, 0, screen_width, screen_height)  # 전체 화면으로 설정

    elif folder_count == 2:
        # 창을 50%, 50%로 분할
        for i, win in enumerate(open_windows):
            hwnd = win._hWnd
            set_window_position(hwnd, i * (screen_width // 2), 0, screen_width // 2, screen_height)

    elif folder_count == 4:
        # 창을 전체 화면 4등분으로 분할
        positions = [
            (0, 0),  # 왼쪽 위
            (screen_width // 2, 0),  # 오른쪽 위
            (0, screen_height // 2),  # 왼쪽 아래
            (screen_width // 2, screen_height // 2)  # 오른쪽 아래
        ]
        for i, win in enumerate(open_windows[:4]):  # 최대 4개만 배치
            hwnd = win._hWnd
            x, y = positions[i]
            set_window_position(hwnd, x, y, screen_width // 2, screen_height // 2)

    else:
        messagebox.showinfo("알림", "현재 열려 있는 파일 탐색기 수가 지원 범위를 초과했습니다 (최대 4개).")
        
def open_tt_folder():
    """
    지정된 4개의 폴더를 열고 화면을 4분할하여 배치합니다.
    """
    # 열고자 하는 폴더 경로 리스트
    folder_paths = [
        r"C:\Radiant Vision Systems Data\TrueTest\Sequence",
        r"C:\Radiant Vision Systems Data\TrueTest\AppData",
        r"C:\Radiant Vision Systems Data\TrueTest\AppData\1.8",
        r"C:\Program Files\Radiant Vision Systems\TrueTest 1.8"
    ]
    time.sleep(1)
    # 폴더 열기 시도
    for path in folder_paths:
        if os.path.exists(path):
            subprocess.Popen(f'explorer "{path}"')
            time.sleep(1)  # 창이 열리기까지 대기
        else:
            messagebox.showerror("오류", f"폴더를 찾을 수 없습니다:\n{path}")
    
    # 열린 창 탐지
    time.sleep(2)  # 모든 창이 열릴 때까지 대기
    windows = gw.getWindowsWithTitle("")
    opened_windows = [win for win in windows if any(folder in win.title for folder in ["Sequence", "AppData", "TrueTest","1.8"])]

    if len(opened_windows) < 4:
        messagebox.showwarning("경고", "모든 폴더 창이 열리지 않았습니다.")

    # 화면 크기 가져오기
    primary_screen = gw.getWindowsWithTitle(gw.getActiveWindow().title)[0]
    screen_width = primary_screen.width
    screen_height = primary_screen.height

    # 4분할 위치 계산
    positions = [
        (0, 0, screen_width // 2, screen_height // 2),  # 좌상단
        (screen_width // 2, 0, screen_width, screen_height // 2),  # 우상단
        (0, screen_height // 2, screen_width // 2, screen_height),  # 좌하단
        (screen_width // 2, screen_height // 2, screen_width, screen_height)  # 우하단
    ]

    # 창 배치
    for win, pos in zip(opened_windows[:4], positions):
        win.resizeTo(pos[2] - pos[0], pos[3] - pos[1])  # 창 크기 조정
        win.moveTo(pos[0], pos[1])  # 창 위치 이동
            
            
def open_download_folder():
    """
    지정된 3개의 폴더를 동시에 엽니다.
    """
    # 열고자 하는 폴더 경로 리스트
    folder_paths = [
        r"C:\Users\ssa2p\Downloads",
        
    ]

    # 폴더 열기 시도
    for path in folder_paths:
        if os.path.exists(path):
            os.startfile(path)
        else:
            messagebox.showerror("오류", f"폴더를 찾을 수 없습니다:\n{path}")


def open_everything_folder():
   # try:
        
        
            everything_path = r"C:\Program Files\Everything\Everything.exe"  
            os.startfile(everything_path)   
           
            
            
def open_google_in_edge():
    """
    Microsoft Edge를 실행하여 https://www.google.co.kr/ 를 엽니다.
    """
    try:
        # Microsoft Edge 실행 파일 경로
        edge_path = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
        if not os.path.exists(edge_path):
            edge_path = r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"

        # Edge 실행 및 URL 열기
        if os.path.exists(edge_path):
            subprocess.Popen([edge_path, "https://www.google.co.kr/","https://chatgpt.com/g/g-FvT4UOsoA-caesgpt/","https://www.naver.com/","https://news.nate.com/"])
            
        else:
            raise FileNotFoundError("Microsoft Edge 실행 파일을 찾을 수 없습니다.")

    except FileNotFoundError as e:
        # Edge 실행 불가능할 때 오류 처리
        messagebox.showerror("오류", str(e))
        
def open_mail_in_edge():
    """
    Microsoft Edge를 실행하여 https://www.google.co.kr/ 를 엽니다.
    """
    try:
        # Microsoft Edge 실행 파일 경로
        edge_path = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
        if not os.path.exists(edge_path):
            edge_path = r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"

        # Edge 실행 및 URL 열기
        if os.path.exists(edge_path):
            subprocess.Popen([edge_path, "https://mail.naver.com/v2/folders/-1","https://mail.google.com/mail/u/0/#inbox"])
            
        else:
            raise FileNotFoundError("Microsoft Edge 실행 파일을 찾을 수 없습니다.")

    except FileNotFoundError as e:
        # Edge 실행 불가능할 때 오류 처리
        messagebox.showerror("오류", str(e))
        

def open_sharepoint():
    
    try:
        # Microsoft Edge 실행 파일 경로
        edge_path = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
        if not os.path.exists(edge_path):
            edge_path = r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"

        # Edge 실행 및 URL 열기
        if os.path.exists(edge_path):
            subprocess.Popen([edge_path, "https://radiantvs.sharepoint.com/sites/RVSKorea/Shared%20Documents/Forms/AllItems.aspx?id=%2Fsites%2FRVSKorea%2FShared%20Documents%2FFile%20share%2FMiguel&p=true/"])
            
        else:
            raise FileNotFoundError("Microsoft Edge 실행 파일을 찾을 수 없습니다.")

    except FileNotFoundError as e:
        # Edge 실행 불가능할 때 오류 처리
        messagebox.showerror("오류", str(e))
        
def create_and_open_excel():
  
    # 경로 설정
    base_path = "F:\\01_RVS\\05_Miguel"
    today_date = datetime.now().strftime("%Y%m%d")
    folder_path = os.path.join(base_path, today_date)
    file_name = f"{today_date}.xlsx"  # 파일 형식을 xlsx로 설정
    file_path = os.path.join(folder_path, file_name)

    # 1. 오늘 날짜의 폴더 만들기
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"폴더 생성됨: {folder_path}")
    else:
        print(f"폴더 이미 존재함: {folder_path}")

    # 2. 오늘 날짜의 빈 엑셀 파일 만들기
    if not os.path.exists(file_path):
        try:
            wb = openpyxl.Workbook()  # 빈 워크북 생성
            wb.save(file_path)  # 파일 저장
            print(f"엑셀 파일 생성됨: {file_path}")
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 파일 생성 중 오류 발생: {e}")
            return
    else:
        print(f"엑셀 파일 이미 존재함: {file_path}")

    # 3. 엑셀 파일 열기
    try:
        os.startfile(file_path)  # 윈도우에서 파일 열기
        print(f"엑셀 파일 열림: {file_path}")
    except Exception as e:
        messagebox.showerror("오류", f"엑셀 파일 열기 실패: {e}")

def open_recent_ppt():
    # Windows "최근 항목" 경로
    recent_folder = os.path.expanduser(r"~\AppData\Roaming\Microsoft\Windows\Recent")
    
    if not os.path.exists(recent_folder):
        messagebox.showerror("오류", f"최근 항목 폴더를 찾을 수 없습니다: {recent_folder}")
        return

    # "최근 항목"에서 PowerPoint 파일만 검색
    recent_files = glob.glob(os.path.join(recent_folder, "*.lnk"))  # .lnk 파일은 바로가기 파일
    ppt_files = [f for f in recent_files if ".ppt" in os.path.basename(f).lower()]

    if not ppt_files:
        messagebox.showinfo("알림", "최근에 열어본 PowerPoint 파일이 없습니다.")
        return

    # 가장 최근 파일 찾기
    recent_file = max(ppt_files, key=os.path.getmtime)
    print(f"가장 최근 파일(바로가기): {recent_file}")

    try:
        # 바로가기를 통해 원본 파일 실행
        os.startfile(recent_file)
        # messagebox.showinfo("성공", f"최근에 열어본 파일을 실행했습니다:\n{os.path.basename(recent_file)}")
    except Exception as e:
         messagebox.showerror("오류", f"파일 실행 중 문제가 발생했습니다: {e}")

def fetch_exchange_rates_from_hana():
   if __name__ == "__main__":
    
        start = str(date.today() - timedelta(days=1825))   # years 매개변수는 없다?? 3년 = 1095, 5년 = 1825
        end = str(date.today() + timedelta(days=1))
        rate_data = yf.download(['USDKRW=X'],start=start, end=end)

        today_rate = rate_data['Close'].iloc[-1]
        today_rate = round(today_rate,2)    # 현재 환율 구하기
        print(today_rate)   # 현재 환율 프린트

        mean_rate = rate_data['Close'].mean()  
        mean_rate = round(mean_rate,2)      # 기간 평균 환율 구하기
        print(mean_rate)


        # if today_rate < mean_rate:
        #     # slackout('현재 환율이 ' + str(today_rate) + '원으로, 환율 5년 평균 ' + str(mean_rate) + '보다 낮습니다.')
        #     print('현재 환율이 ' + str(today_rate) + '원으로, 환율 5년 평균 ' + str(mean_rate) + '원 보다 낮습니다.')
        # elif (mean_rate * 1.1) > today_rate:
        #     # slackout('현재 환율이 ' + str(today_rate) + '원으로, 환율 5년 평균 ' + str(mean_rate) + '보다 10% 이상 높습니다.')
        #     print('현재 환율이 ' + str(today_rate) + '원으로, 환율 5년 평균 ' + str(mean_rate) + '원 보다 10% 이상 높습니다.')

def on_closing():
    """창 닫기 버튼 이벤트: 창을 숨기고 시스템 트레이로 이동"""
    root.withdraw()  # Tkinter 창 숨기기
    show_tray_icon()  # 시스템 트레이 아이콘 표시


def show_tray_icon():
    """시스템 트레이 아이콘 표시"""
    def quit_app(icon, item):
        """프로그램 종료"""
        icon.stop()  # 트레이 아이콘 중지
        root.destroy()  # Tkinter 종료

    def show_window(icon, item):
        """숨겨진 Tkinter 창 다시 표시"""
        icon.stop()  # 트레이 아이콘 중지
        root.deiconify()  # Tkinter 창 다시 표시

    # 트레이 아이콘 이미지 설정
    
    tray_image = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\tray.png")  # 적절한 아이콘 경로
   

    # 트레이 메뉴 설정
    menu = Menu(
        MenuItem("Show", show_window),
        MenuItem("Exit", quit_app)
    )

    # 트레이 아이콘 생성
    icon = Icon("AppName", tray_image, menu=menu)
     
    
    # 트레이 아이콘 실행 (별도 스레드에서 실행)
    threading.Thread(target=icon.run, daemon=True).start()
    


 
def update_usage(system_label):
    while True:
        # CPU 사용량
        cpu_usage = psutil.cpu_percent(interval=1)

        # 메모리 사용량
        memory_info = psutil.virtual_memory()
        memory_usage_percent = memory_info.percent
        memory_used_mb = memory_info.used // (1024 * 1024)  # MB 단위로 변환
        memory_total_mb = memory_info.total // (1024 * 1024)  # MB 단위로 변환

        # 하나의 라벨로 업데이트
        system_label.config(
            text=(
                f"CPU: {cpu_usage}%  "
                f"Memory: {memory_usage_percent}% "
                f"({memory_used_mb}MB / {memory_total_mb}MB)"
            )
        )

        time.sleep(1)  # 1초 대기


def open_calculator():
    try:
        # Windows 계산기 실행
        subprocess.Popen('calc.exe')
    except Exception as e:
        print(f"Error: {e}")

def close_window():
    root.destroy()  # Tkinter 창 종료
    



# TortoiseGit 실행 파일 경로
TOROISEGIT_PATH = r"C:\Program Files\TortoiseGit\bin\TortoiseGitProc.exe"


# 로컬 저장소 경로
LOCAL_REPO_PATH1 = r"F:\01_RVS\00_RadiantDevGIT\releaseSDC-Tablet_Release_20231120\SDC"
LOCAL_REPO_PATH2= r"F:\01_RVS\00_RadiantDevGIT\releaseLGD-Tablet_Release_20231120\LG"
LOCAL_REPO_PATH3 = r"F:\01_RVS\00_RadiantDevGIT\releaseLGD-Mobile_Release_20231120\LG"

# 원격 및 브랜치 정보
REMOTE_NAME = "origin"
REMOTE_BRANCH1 = "feature/SDC-Tablet_Release_20231120"
REMOTE_BRANCH2 = "feature/LGD-Tablet_Release_20231120"
REMOTE_BRANCH3 = "feature/LGD-Mobile_Release_20231120"

def pull_with_tortoisegit():
    # try:
    #     # TortoiseGit 실행 파일 경로 확인
    #     if not os.path.exists(TOROISEGIT_PATH):
    #         print(f"TortoiseGit 실행 파일을 찾을 수 없습니다: {TOROISEGIT_PATH}")
    #         return
        
    #     # 로컬 Git 저장소 경로 확인
    #     if not os.path.exists(LOCAL_REPO_PATH1):
    #         print(f"로컬 Git 저장소 경로를 찾을 수 없습니다: {LOCAL_REPO_PATH1}")
    #         return

        # TortoiseGit Pull 명령 실행
        subprocess.run([
            TOROISEGIT_PATH,
            "/command:pull",
            f"/path:{LOCAL_REPO_PATH1}",
            f"/url:{REMOTE_NAME}/{REMOTE_BRANCH1}",
            "/closeonend:1"
        ], check=True)
        
        subprocess.run([
            TOROISEGIT_PATH,
            "/command:pull",
            f"/path:{LOCAL_REPO_PATH2}",
            f"/url:{REMOTE_NAME}/{REMOTE_BRANCH2}",
            "/closeonend:1"
        ], check=True)
        
        subprocess.run([
            TOROISEGIT_PATH,
            "/command:pull",
            f"/path:{LOCAL_REPO_PATH3}",
            f"/url:{REMOTE_NAME}/{REMOTE_BRANCH3}",
            "/closeonend:1"
        ], check=True)
        
        
    #     print("Git pull 작업이 완료되었습니다!")

    # except subprocess.CalledProcessError as e:
    #     print(f"Git pull 중 오류 발생: {e}")
    # except Exception as e:
    #     print(f"알 수 없는 오류 발생: {e}")
        
    # try:
    #     # TortoiseGit 실행 파일 경로 확인
    #     if not os.path.exists(TOROISEGIT_PATH):
    #         print(f"TortoiseGit 실행 파일을 찾을 수 없습니다: {TOROISEGIT_PATH}")
    #         return
        
    #     # 로컬 Git 저장소 경로 확인
    #     if not os.path.exists(LOCAL_REPO_PATH2):
    #         print(f"로컬 Git 저장소 경로를 찾을 수 없습니다: {LOCAL_REPO_PATH2}")
    #         return

    #     # TortoiseGit Pull 명령 실행
    #     subprocess.run([
    #         TOROISEGIT_PATH,
    #         "/command:pull",
    #         f"/path:{LOCAL_REPO_PATH2}",
    #         f"/url:{REMOTE_NAME}/{REMOTE_BRANCH2}",
    #         "/closeonend:1"
    #     ], check=True)
    #     print("Git pull 작업이 완료되었습니다!")

    # except subprocess.CalledProcessError as e:
    #     print(f"Git pull 중 오류 발생: {e}")
    # except Exception as e:
    #     print(f"알 수 없는 오류 발생: {e}")
    
    
    # try:
    #     # TortoiseGit 실행 파일 경로 확인
    #     if not os.path.exists(TOROISEGIT_PATH):
    #         print(f"TortoiseGit 실행 파일을 찾을 수 없습니다: {TOROISEGIT_PATH}")
    #         return
        
    #     # 로컬 Git 저장소 경로 확인
    #     if not os.path.exists(LOCAL_REPO_PATH3):
    #         print(f"로컬 Git 저장소 경로를 찾을 수 없습니다: {LOCAL_REPO_PATH3}")
    #         return

    #     # TortoiseGit Pull 명령 실행
    #     subprocess.run([
    #         TOROISEGIT_PATH,
    #         "/command:pull",
    #         f"/path:{LOCAL_REPO_PATH3}",
    #         f"/url:{REMOTE_NAME}/{REMOTE_BRANCH3}",
    #         "/closeonend:1"
    #     ], check=True)
    #     print("Git pull 작업이 완료되었습니다!")

    # except subprocess.CalledProcessError as e:
    #     print(f"Git pull 중 오류 발생: {e}")
    # except Exception as e:
    #     print(f"알 수 없는 오류 발생: {e}")
        
def create_gui():
    # Tkinter 윈도우 생성
    global root
    root = tk.Tk()
    root.title("Automation work")
    
    # 상태 표시줄 아이콘 변경
    icon_path = r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico.ico"  # .ico 파일 경로
    try:
        root.iconbitmap(icon_path)
    except Exception as e:
        print(f"아이콘 로드 실패: {e}")

    # 화면 크기와 위치 계산
    window_width = 380
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()  # 전체 화면 높이

    # 창 높이를 화면 높이로 설정
    window_height = screen_height

    # 창 위치 계산 (오른쪽 끝 상단)
    x_pos = screen_width - window_width  # 오른쪽 끝
    y_pos = 0  # 상단

    # 창 크기 및 위치 설정
    root.geometry(f"{window_width}x{window_height}+{x_pos}+{y_pos}")
    root.configure(bg="#121212")  # 배경색: 어두운 색상

    # 현재 시간 표시 라벨
    title_label = ttk.Label(root, font=("Arial", 20, "bold"), background="#121212", foreground="#FFFFFF")
    title_label.pack()  # padding 제거로 간격 최소화

    # 시간 업데이트 시작
    update_time(title_label, root)

    # 현재 달력 표시
    CalendarApp(root)
    
    
    
    # CPU와 메모리 사용량 통합 라벨
    system_label = ttk.Label(
        root,
        text="Calculating...",
        font=("Arial", 12, "bold"),
        background="#121212",
        foreground="#00FF00"  # 초록색 글씨
    )
    system_label.pack(pady=0)

    # 사용량 업데이트를 별도 스레드에서 실행
    threading.Thread(target=update_usage, args=(system_label,), daemon=True).start()

    # 버튼 프레임 생성
    button_frame = ttk.Frame(root, padding="0")
    button_frame.pack(side="top", pady=1, anchor="w")
    
    # 스타일 변경 (다크 모드 스타일 설정)
    style = ttk.Style()
    style.theme_use("clam")  # 스타일 테마 설정
    style.configure(
        "TButton",
        background="#333333",  # 버튼 배경색
        foreground="#FFFFFF",  # 버튼 글씨색
        borderwidth=0,
        padding=5,
        focuscolor="none"
    )
    style.map(
        "TButton",
        background=[("active", "#555555")],  # 활성화 시 색상
        foreground=[("active", "#FFFFFF")]
    )
    style.configure(
        "TLabel",
        background="#121212",  # 라벨 배경색
        foreground="#FFFFFF"  # 라벨 글씨색
    )

    # 아이콘 불러오기
    tortoise_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\tortoise.png").resize((32, 32), Image.LANCZOS)
    tortoise_icon_tk = ImageTk.PhotoImage(tortoise_icon)

    file_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\file.png").resize((32, 32), Image.LANCZOS)
    file_icon_tk = ImageTk.PhotoImage(file_icon)

    python_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\python.png").resize((32, 32), Image.LANCZOS)
    python_icon_tk = ImageTk.PhotoImage(python_icon)

    outlook_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\outlook_icon.png").resize((32, 32), Image.LANCZOS)
    outlook_icon_tk = ImageTk.PhotoImage(outlook_icon)

    onenote_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\onenote_icon.png").resize((32, 32), Image.LANCZOS)
    onenote_icon_tk = ImageTk.PhotoImage(onenote_icon)
    
    teams_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\teams.png").resize((32, 32), Image.LANCZOS)
    teams_icon_tk = ImageTk.PhotoImage(teams_icon)
    
    kakao_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\kakao.png").resize((32, 32), Image.LANCZOS)
    kakao_icon_tk = ImageTk.PhotoImage(kakao_icon)
    
    tt_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\tt.png").resize((32, 32), Image.LANCZOS)
    tt_icon_tk = ImageTk.PhotoImage(tt_icon)
    
    closett_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\exit.png").resize((32, 32), Image.LANCZOS)
    closett_icon_tk = ImageTk.PhotoImage(closett_icon)

    imagej_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\imagej.png").resize((32, 32), Image.LANCZOS)
    imagej_icon_tk = ImageTk.PhotoImage(imagej_icon)
    
    createfolder_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\createfolder.png").resize((32, 32), Image.LANCZOS)
    createfolder_icon_tk = ImageTk.PhotoImage(createfolder_icon)
    
    report_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\report.png").resize((32, 32), Image.LANCZOS)
    report_icon_tk = ImageTk.PhotoImage(report_icon)
    
    send_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\send.png").resize((32, 32), Image.LANCZOS)
    send_icon_tk = ImageTk.PhotoImage(send_icon)
    
    tt_folder_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\tt_folder.png").resize((32, 32), Image.LANCZOS)
    tt_folder_icon_tk = ImageTk.PhotoImage(tt_folder_icon)
    
    download_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\download.png").resize((32, 32), Image.LANCZOS)
    download_icon_tk = ImageTk.PhotoImage(download_icon)
    
    edge_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\edge.png").resize((32, 32), Image.LANCZOS)
    edge_icon_tk = ImageTk.PhotoImage(edge_icon)
    
    mail_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\mail.png").resize((32, 32), Image.LANCZOS)
    mail_icon_tk = ImageTk.PhotoImage(mail_icon)
    
    note_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\note.png").resize((32, 32), Image.LANCZOS)
    note_icon_tk = ImageTk.PhotoImage(note_icon)
    
    everything_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\Everything.png").resize((32, 32), Image.LANCZOS)
    everything_icon_tk = ImageTk.PhotoImage(everything_icon)
    
    winmerge_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\winmerge.png").resize((32, 32), Image.LANCZOS)
    winmerge_icon_tk = ImageTk.PhotoImage(winmerge_icon)
    
    excel_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\excel.png").resize((32, 32), Image.LANCZOS)
    excel_icon_tk = ImageTk.PhotoImage(excel_icon)
    
    ppt_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\ppt.png").resize((32, 32), Image.LANCZOS)
    ppt_icon_tk = ImageTk.PhotoImage(ppt_icon)
    
    sharepoint_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\sharepoint.png").resize((32, 32), Image.LANCZOS)
    sharepoint_icon_tk = ImageTk.PhotoImage(sharepoint_icon)
    
    cal_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\cal.png").resize((32, 32), Image.LANCZOS)
    cal_icon_tk = ImageTk.PhotoImage(cal_icon)
    
    shutdown_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\shutdown.png").resize((32, 32), Image.LANCZOS)
    shutdown_icon_tk = ImageTk.PhotoImage(shutdown_icon)
    
    turnoff_icon = Image.open(r"F:\OneDrive - Radiant Vision Systems\02_Miguel_Toy\01_Python\업무효율화\ico\turnoff.png").resize((32, 32), Image.LANCZOS)
    turnoff_icon_tk = ImageTk.PhotoImage(turnoff_icon)


    # 버튼 프레임 생성 및 상단 정렬
    button_frame = ttk.Frame(root, padding="1", style="Dark.TFrame")
    button_frame.pack(side="top", pady=20, anchor="w")
    # 스타일 설정 (프레임의 배경을 어둡게 설정)
    style.configure("Dark.TFrame", background="#121212")  # 검은색 배경

    # grid 레이아웃 비율 설정 (열 분배)
    for col in range(8):  # 열은 6개로 설정
        button_frame.grid_columnconfigure(col, weight=1)

   # 버튼 추가
    btn0 = ttk.Button(button_frame, image=tortoise_icon_tk, command=pull_with_tortoisegit)
    btn0.image = tortoise_icon_tk  # 가비지 컬렉션 방지
    btn0.grid(row=0, column=0, padx=2, pady=2)  # 간격 줄이기

    btn00 = ttk.Button(button_frame, image=python_icon_tk, command=make_python)
    btn00.image = python_icon_tk  # 가비지 컬렉션 방지
    btn00.grid(row=0, column=1, padx=2, pady=2)

    btn1 = ttk.Button(button_frame, image=outlook_icon_tk, command=open_outlook)
    btn1.image = outlook_icon_tk  # 가비지 컬렉션 방지
    btn1.grid(row=1, column=0, padx=2, pady=2)

    btn2 = ttk.Button(button_frame, image=onenote_icon_tk, command=open_onenote)
    btn2.image = onenote_icon_tk  # 이미지 참조 유지
    btn2.grid(row=1, column=1, padx=2, pady=2)

    btn3 = ttk.Button(button_frame, image=teams_icon_tk, command=open_teams)
    btn3.image = teams_icon_tk  # 가비지 컬렉션 방지
    btn3.grid(row=1, column=2, padx=2, pady=2)

    btn4 = ttk.Button(button_frame, image=kakao_icon_tk, command=open_kt)
    btn4.image = kakao_icon_tk  # 가비지 컬렉션 방지
    btn4.grid(row=1, column=3, padx=2, pady=2)
    
    btn5 = ttk.Button(button_frame, image=sharepoint_icon_tk, command=open_sharepoint)
    btn5.image = sharepoint_icon_tk  # 가비지 컬렉션 방지
    btn5.grid(row=1, column=4, padx=2, pady=2)
    
    btn17 = ttk.Button(button_frame, image=cal_icon_tk, command=open_calculator)
    btn17.image = cal_icon_tk  # 가비지 컬렉션 방지
    btn17.grid(row=1, column=10, padx=2, pady=2)

    btn21 = ttk.Button(button_frame, image=tt_icon_tk, command=open_truetest)
    btn21.image = tt_icon_tk  # 가비지 컬렉션 방지
    btn21.grid(row=2, column=0, padx=2, pady=2)
    
    btn22 = ttk.Button(button_frame, image=tt_folder_icon_tk, command=open_tt_folder)
    btn22.image = tt_folder_icon_tk  # 가비지 컬렉션 방지
    btn22.grid(row=2, column=1, padx=2, pady=2)

    btn26 = ttk.Button(button_frame, image=imagej_icon_tk, command=open_imagej)
    btn26.image = imagej_icon_tk  # 가비지 컬렉션 방지
    btn26.grid(row=2, column=5, padx=2, pady=2, sticky="w")

    btn27 = ttk.Button(button_frame, image=closett_icon_tk, command=close_truetest)
    btn27.image = tt_icon_tk
    btn27.grid(row=2, column=6, padx=2, pady=2, sticky="w")
    
    btn27 = ttk.Button(button_frame, image=closett_icon_tk, command=close_truetest)
    btn27.image = tt_icon_tk
    btn27.grid(row=2, column=10, padx=2, pady=2, sticky="w")
    
    # btn101을 10번째 행(9번째 인덱스)에 배치
    btn41 = ttk.Button(button_frame, image=download_icon_tk, command=open_download_folder)
    btn41.image = download_icon_tk  # 가비지 컬렉션 방지
    btn41.grid(row=4, column=0, padx=2, pady=2)
    
    btn42 = ttk.Button(button_frame, image=everything_icon_tk, command=open_everything_folder)
    btn42.image = everything_icon_tk  # 가비지 컬렉션 방지
    btn42.grid(row=4, column=1, padx=2, pady=2)
    
    btn43 = ttk.Button(button_frame, image=winmerge_icon_tk, command=open_winmerge)
    btn43.image = winmerge_icon_tk  # 가비지 컬렉션 방지
    btn43.grid(row=4, column=2, padx=2, pady=2)
    
    btn51 = ttk.Button(button_frame, image=edge_icon_tk, command=open_google_in_edge)
    btn51.image = edge_icon_tk  # 가비지 컬렉션 방지
    btn51.grid(row=5, column=0, padx=2, pady=2)
    
    btn52 = ttk.Button(button_frame, image=mail_icon_tk, command=open_mail_in_edge)
    btn52.image = mail_icon_tk  # 가비지 컬렉션 방지
    btn52.grid(row=5, column=1, padx=2, pady=2)
    
    
    
    btn101 = ttk.Button(button_frame, image=tt_icon_tk, command=open_truetest)
    btn101.image = tt_icon_tk  # 가비지 컬렉션 방지
    btn101.grid(row=6, column=0, padx=2, pady=2)
    
    btn101 = ttk.Button(button_frame, image=tt_icon_tk, command=open_truetest)
    btn101.image = tt_icon_tk  # 가비지 컬렉션 방지
    btn101.grid(row=7, column=0, padx=2, pady=2)
    
    btn801 = ttk.Button(button_frame, image=tt_icon_tk, command=fetch_exchange_rates_from_hana)
    btn801.image = tt_icon_tk  # 가비지 컬렉션 방지
    btn801.grid(row=8, column=0, padx=2, pady=2)
    
    btn91 = ttk.Button(button_frame, image=note_icon_tk, command=make_folder)
    btn91.image = note_icon_tk  # 가비지 컬렉션 방지
    btn91.grid(row=9, column=0, padx=2, pady=2)
    
    btn92 = ttk.Button(button_frame, image=excel_icon_tk, command=create_and_open_excel)
    btn92.image = excel_icon_tk  # 가비지 컬렉션 방지
    btn92.grid(row=9, column=1, padx=2, pady=2)
    
    btn93 = ttk.Button(button_frame, image=createfolder_icon_tk, command=arrange_folders)
    btn93.image = createfolder_icon_tk  # 가비지 컬렉션 방지
    btn93.grid(row=9, column=2, padx=2, pady=2)
    
    btn94 = ttk.Button(button_frame, image=createfolder_icon_tk, command=lambda: create_folders(root))
    btn94.image = createfolder_icon_tk  # 가비지 컬렉션 방지
    btn94.grid(row=9, column=3, padx=2, pady=2)
    
    btn95 = ttk.Button(button_frame, image=ppt_icon_tk, command=open_recent_ppt)
    btn95.image = ppt_icon_tk  # 가비지 컬렉션 방지
    btn95.grid(row=9, column=4, padx=2, pady=2)
    
    
    btn101 = ttk.Button(button_frame, image=report_icon_tk, command=copy_both_reports)
    btn101.image = report_icon_tk  # 가비지 컬렉션 방지
    btn101.grid(row=10, column=0, padx=2, pady=2)
    
    btn102 = ttk.Button(button_frame, image=send_icon_tk, command=open_outlook_and_new_mail)
    btn102.image = send_icon_tk  # 가비지 컬렉션 방지
    btn102.grid(row=10, column=1, padx=2, pady=2)
    
    btn1010 = ttk.Button(button_frame, image=turnoff_icon_tk, command=close_window)
    btn1010.image = turnoff_icon_tk  # 가비지 컬렉션 방지
    btn1010.grid(row=10, column=10, padx=2, pady=2)
    
    # btn1110 = ttk.Button(button_frame, image=shutdown_icon_tk, command=lock_screen)
    # btn1110.image = shutdown_icon_tk  # 가비지 컬렉션 방지
    # btn1110.grid(row=11, column=10, padx=2, pady=2)
    
    
    
    
    
    # Tkinter 메인 루프 실행
    root.protocol("WM_DELETE_WINDOW", on_closing)  # 창 닫기 이벤트 재정의
    root.mainloop()

# GUI 실행
create_gui()
